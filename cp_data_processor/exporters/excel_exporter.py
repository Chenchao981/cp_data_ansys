import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from ..data_models.cp_data import CPParameter, CPWafer, CPLot

class ExcelExporter:
    """用于将CP数据和分析结果导出到Excel文件的导出器"""
    
    def __init__(self, workbook=None):
        """
        初始化Excel导出器
        
        Args:
            workbook: 现有的openpyxl工作簿，默认为None创建新的工作簿
        """
        self.workbook = workbook if workbook else Workbook()
        # 如果是新创建的工作簿，重命名默认工作表
        if len(self.workbook.worksheets) == 1 and self.workbook.worksheets[0].title == "Sheet":
            self.workbook.worksheets[0].title = "Summary"
    
    def add_dataframe(self, df, sheet_name, start_cell='A1', include_index=False, 
                     auto_fit=True, apply_styles=True):
        """
        添加DataFrame到指定工作表
        
        Args:
            df: 要添加的DataFrame
            sheet_name: 工作表名称，不存在则创建
            start_cell: 起始单元格，默认为'A1'
            include_index: 是否包含索引，默认为False
            auto_fit: 是否自动调整列宽，默认为True
            apply_styles: 是否应用样式，默认为True
        
        Returns:
            添加数据的工作表
        """
        # 获取或创建工作表
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.create_sheet(sheet_name)
        
        # 解析起始单元格
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        coordinate = coordinate_from_string(start_cell)
        start_row = coordinate[1]
        start_col = column_index_from_string(coordinate[0])
        
        # 添加数据
        rows = dataframe_to_rows(df, index=include_index, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=start_row + r_idx - 1, column=start_col + c_idx - 1, value=value)
        
        # 应用样式
        if apply_styles:
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # 表头行
            for c_idx in range(len(df.columns) + (1 if include_index else 0)):
                cell = sheet.cell(row=start_row, column=start_col + c_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 数据行
            data_alignment = Alignment(horizontal='center', vertical='center')
            for r_idx in range(len(df)):
                for c_idx in range(len(df.columns) + (1 if include_index else 0)):
                    cell = sheet.cell(row=start_row + r_idx + 1, column=start_col + c_idx)
                    cell.alignment = data_alignment
                    cell.border = thin_border
        
        # 自动调整列宽
        if auto_fit:
            for c_idx in range(len(df.columns) + (1 if include_index else 0)):
                column = get_column_letter(start_col + c_idx)
                max_length = 0
                for r_idx in range(len(df) + 1):  # +1 for header
                    cell = sheet.cell(row=start_row + r_idx, column=start_col + c_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width
        
        return sheet
    
    def add_figure(self, fig, sheet_name, position='A1', width=None, height=None):
        """
        添加matplotlib图表到指定工作表
        
        Args:
            fig: matplotlib图表对象
            sheet_name: 工作表名称，不存在则创建
            position: 插入位置，默认为'A1'
            width: 图像宽度，默认为None使用原始大小
            height: 图像高度，默认为None使用原始大小
        
        Returns:
            添加图表的工作表
        """
        # 获取或创建工作表
        if sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.create_sheet(sheet_name)
        
        # 将图表保存到内存中
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120)
        buf.seek(0)
        
        # 创建图像对象
        img = Image(buf)
        
        # 调整大小
        if width and height:
            img.width = width
            img.height = height
        
        # 添加图像到工作表
        sheet.add_image(img, position)
        
        return sheet
    
    def add_cp_lot(self, cp_lot, include_summary=True, include_wafers=True, include_parameters=True):
        """
        添加CPLot对象到工作簿
        
        Args:
            cp_lot: CPLot对象
            include_summary: 是否包含摘要信息，默认为True
            include_wafers: 是否包含晶圆信息，默认为True
            include_parameters: 是否包含参数信息，默认为True
        
        Returns:
            self，支持链式调用
        """
        # 添加摘要信息
        if include_summary:
            summary_sheet = self._get_or_create_sheet("Summary")
            summary_info = {
                "批次号": cp_lot.lot_id,
                "总晶圆数": len(cp_lot.wafers),
                "总Die数": sum(len(wafer.dies) for wafer in cp_lot.wafers),
                "参数数量": len(cp_lot.parameters)
            }
            
            # 添加到工作表
            row = 1
            for key, value in summary_info.items():
                summary_sheet.cell(row=row, column=1, value=key)
                summary_sheet.cell(row=row, column=2, value=value)
                row += 1
            
            # 设置样式
            for r in range(1, row):
                summary_sheet.cell(row=r, column=1).font = Font(bold=True)
        
        # 添加晶圆信息
        if include_wafers and cp_lot.wafers:
            wafer_data = []
            for wafer in cp_lot.wafers:
                good_dies = sum(1 for die in wafer.dies if die.bin == 1)
                wafer_data.append({
                    "晶圆ID": wafer.wafer_id,
                    "总Die数": len(wafer.dies),
                    "良品Die数": good_dies,
                    "良率(%)": (good_dies / len(wafer.dies) * 100) if len(wafer.dies) > 0 else 0
                })
            
            if wafer_data:
                wafer_df = pd.DataFrame(wafer_data)
                self.add_dataframe(wafer_df, "Wafers")
        
        # 添加参数信息
        if include_parameters and cp_lot.parameters:
            param_data = []
            for param in cp_lot.parameters:
                param_data.append({
                    "参数名": param.name,
                    "单位": param.unit,
                    "下限": param.lsl,
                    "上限": param.usl,
                    "描述": param.description
                })
            
            if param_data:
                param_df = pd.DataFrame(param_data)
                self.add_dataframe(param_df, "Parameters")
        
        return self
    
    def add_analysis_results(self, results, sheet_name="Analysis", start_cell='A1'):
        """
        添加分析结果到工作簿
        
        Args:
            results: 分析结果字典或DataFrame
            sheet_name: 工作表名称，默认为"Analysis"
            start_cell: 起始单元格，默认为'A1'
        
        Returns:
            self，支持链式调用
        """
        # 将结果转换为DataFrame（如果不是）
        if not isinstance(results, pd.DataFrame):
            # 尝试转换字典为DataFrame
            try:
                results_df = pd.DataFrame.from_dict(results, orient='index')
                # 如果结果层次太深，尝试扁平化
                if isinstance(results_df.iloc[0, 0], dict):
                    flattened_data = []
                    for k1, v1 in results.items():
                        for k2, v2 in v1.items():
                            row_data = {"Category": k1, "Item": k2}
                            if isinstance(v2, dict):
                                row_data.update(v2)
                            else:
                                row_data["Value"] = v2
                            flattened_data.append(row_data)
                    results_df = pd.DataFrame(flattened_data)
            except:
                # 如果转换失败，创建简单的键值对DataFrame
                results_df = pd.DataFrame({"Item": list(results.keys()), "Value": list(results.values())})
        else:
            results_df = results
        
        # 添加到工作表
        self.add_dataframe(results_df, sheet_name, start_cell=start_cell)
        
        return self
    
    def save(self, file_path):
        """
        保存工作簿到文件
        
        Args:
            file_path: 保存路径
        
        Returns:
            保存是否成功的布尔值
        """
        try:
            self.workbook.save(file_path)
            return True
        except Exception as e:
            print(f"保存Excel文件失败: {e}")
            return False
    
    def _get_or_create_sheet(self, sheet_name):
        """获取现有工作表或创建新工作表"""
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            return self.workbook.create_sheet(sheet_name) 